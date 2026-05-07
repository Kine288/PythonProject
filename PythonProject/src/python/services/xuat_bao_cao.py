import tempfile
from pathlib import Path
from typing import Dict, List, Optional

from config.db_config import get_database_connection


def _conn_or_raise():
	conn = get_database_connection()
	if not conn:
		raise ConnectionError("Khong ket noi duoc database")
	return conn


def _temp_file(suffix: str) -> str:
	output_dir = Path(tempfile.gettempdir()) / "python_project_reports"
	output_dir.mkdir(parents=True, exist_ok=True)
	handle = tempfile.NamedTemporaryFile(delete=False, suffix=suffix, dir=output_dir)
	handle.close()
	return handle.name


def lay_ds_canh_bao(hoc_ky_id: str, muc_canh_bao: Optional[int] = None) -> List[Dict]:
	conn = _conn_or_raise()
	try:
		with conn.cursor() as cursor:
			sql = """
				SELECT sv.msv, sv.ho_ten, lsh.ten_lop,
				       kq.gpa_tich_luy_he4, kq.xep_loai, kq.muc_canh_bao
				FROM ket_qua_hoc_ky kq
				JOIN sinh_vien sv ON sv.sinh_vien_id = kq.sinh_vien_id
				LEFT JOIN lop_sinh_hoat lsh ON lsh.lop_id = sv.lop_id
				WHERE kq.hoc_ky_id = %s AND kq.muc_canh_bao > 0
			"""
			params = [hoc_ky_id]
			if muc_canh_bao is not None:
				sql += " AND kq.muc_canh_bao = %s"
				params.append(muc_canh_bao)
			sql += " ORDER BY kq.muc_canh_bao DESC, kq.gpa_tich_luy_he4 ASC, sv.msv ASC"

			cursor.execute(sql, tuple(params))
			return cursor.fetchall()
	finally:
		conn.close()


def thong_ke_xep_loai(hoc_ky_id: str) -> List[Dict]:
	conn = _conn_or_raise()
	try:
		with conn.cursor() as cursor:
			cursor.execute(
				"""
				SELECT COALESCE(kq.xep_loai, 'Chua xep loai') AS xep_loai,
				       COUNT(*) AS so_luong
				FROM ket_qua_hoc_ky kq
				WHERE kq.hoc_ky_id = %s
				GROUP BY COALESCE(kq.xep_loai, 'Chua xep loai')
				""",
				(hoc_ky_id,),
			)
			rows = cursor.fetchall()
			tong = sum(int(row["so_luong"]) for row in rows)
			for row in rows:
				row["ty_le"] = round((int(row["so_luong"]) / tong) * 100, 2) if tong else 0
			return rows
	finally:
		conn.close()


def xuat_excel_canh_bao(hoc_ky_id: str, muc_canh_bao: Optional[int] = None) -> str:
	try:
		from openpyxl import Workbook
	except Exception as exc:  # pragma: no cover
		raise RuntimeError("Thieu thu vien openpyxl. Vui long cai dat openpyxl.") from exc

	data = lay_ds_canh_bao(hoc_ky_id, muc_canh_bao)
	file_path = _temp_file(".xlsx")

	wb = Workbook()
	ws = wb.active
	ws.title = "CanhBaoHocVu"
	ws.append(["MSV", "Ten sinh vien", "Lop", "GPA tich luy he 4", "Xep loai", "Muc canh bao"])

	for row in data:
		ws.append(
			[
				row.get("msv"),
				row.get("ho_ten"),
				row.get("ten_lop"),
				row.get("gpa_tich_luy_he4"),
				row.get("xep_loai"),
				row.get("muc_canh_bao"),
			]
		)

	wb.save(file_path)
	return file_path


def xuat_excel_tong_ket_hoc_vu(hoc_ky_id: str, lop_id: Optional[str] = None) -> str:
	try:
		from openpyxl import Workbook
	except Exception as exc:  # pragma: no cover
		raise RuntimeError("Thieu thu vien openpyxl. Vui long cai dat openpyxl.") from exc

	conn = _conn_or_raise()
	try:
		with conn.cursor() as cursor:
			cursor.execute(
				"""
				SELECT
					lsh.lop_id,
					lsh.ma_lop,
					lsh.ten_lop,
					nk.ten_nien_khoa,
					COUNT(sv.sinh_vien_id) AS tong_sv,
					SUM(CASE WHEN kq.kqhk_id IS NOT NULL THEN 1 ELSE 0 END) AS da_tong_ket,
					SUM(CASE WHEN COALESCE(kq.muc_canh_bao, 0) > 0 THEN 1 ELSE 0 END) AS so_sv_canh_bao,
					ROUND(AVG(kq.gpa_tich_luy_he4), 2) AS gpa_tb,
					CASE
						WHEN COUNT(sv.sinh_vien_id) = 0 THEN 'CHUA_CO_DU_LIEU'
						WHEN SUM(CASE WHEN kq.kqhk_id IS NOT NULL THEN 1 ELSE 0 END) < COUNT(sv.sinh_vien_id) THEN 'DANG_XU_LY'
						ELSE 'HOAN_TAT'
					END AS trang_thai_tong_ket
				FROM lop_sinh_hoat lsh
				LEFT JOIN nien_khoa nk ON nk.nien_khoa_id = lsh.nien_khoa_id
				LEFT JOIN sinh_vien sv ON sv.lop_id = lsh.lop_id
				LEFT JOIN ket_qua_hoc_ky kq ON kq.sinh_vien_id = sv.sinh_vien_id AND kq.hoc_ky_id = %s
				GROUP BY lsh.lop_id, lsh.ma_lop, lsh.ten_lop, nk.ten_nien_khoa
				ORDER BY lsh.ma_lop ASC
				""",
				(hoc_ky_id,),
			)
			lop_summary = cursor.fetchall()

			chi_tiet_lop = []
			if lop_id:
				cursor.execute(
					"""
					SELECT
						sv.sinh_vien_id,
						sv.msv,
						sv.ho_ten,
						sv.trang_thai,
						kq.gpa_hk_he4,
						kq.gpa_tich_luy_he4,
						kq.xep_loai,
						kq.muc_canh_bao
					FROM sinh_vien sv
					LEFT JOIN ket_qua_hoc_ky kq ON kq.sinh_vien_id = sv.sinh_vien_id AND kq.hoc_ky_id = %s
					WHERE sv.lop_id = %s
					ORDER BY sv.msv ASC
					""",
					(hoc_ky_id, lop_id),
				)
				chi_tiet_lop = cursor.fetchall()
	finally:
		conn.close()

	file_path = _temp_file(".xlsx")
	wb = Workbook()
	ws_summary = wb.active
	ws_summary.title = "TongKetTheoLop"
	ws_summary.append(["Hoc ky", hoc_ky_id])
	ws_summary.append([])
	ws_summary.append(["Ma lop", "Ten lop", "Khoa", "Tong SV", "Da tong ket", "SV canh bao", "GPA TB", "Trang thai"])

	for row in lop_summary:
		ws_summary.append(
			[
				row.get("ma_lop"),
				row.get("ten_lop"),
				row.get("ten_nien_khoa"),
				row.get("tong_sv"),
				row.get("da_tong_ket"),
				row.get("so_sv_canh_bao"),
				row.get("gpa_tb"),
				row.get("trang_thai_tong_ket"),
			]
		)

	if chi_tiet_lop:
		ws_detail = wb.create_sheet("ChiTietLop")
		ws_detail.append(["MSSV", "Ho ten", "GPA hoc ky (he 4)", "GPA tich luy (he 4)", "Xep loai", "Muc canh bao", "Trang thai hoc tap"])
		for row in chi_tiet_lop:
			ws_detail.append(
				[
					row.get("msv"),
					row.get("ho_ten"),
					row.get("gpa_hk_he4"),
					row.get("gpa_tich_luy_he4"),
					row.get("xep_loai"),
					row.get("muc_canh_bao"),
					row.get("trang_thai"),
				]
			)

	wb.save(file_path)
	return file_path


def _lay_bang_diem_sinh_vien(sinh_vien_id: str):
	conn = _conn_or_raise()
	try:
		with conn.cursor() as cursor:
			cursor.execute(
				"""
				SELECT sv.msv, sv.ho_ten,
				       hk.ten_hoc_ky, mh.ma_mon, mh.ten_mon, mh.so_tin_chi,
				       ds.diem_tong,
				       kq.gpa_hk_he4, kq.gpa_tich_luy_he4, kq.xep_loai
				FROM ds_lhp ds
				JOIN sinh_vien sv ON sv.sinh_vien_id = ds.sinh_vien_id
				JOIN lop_hoc_phan lhp ON lhp.lhp_id = ds.lhp_id
				JOIN mon_hoc mh ON mh.mon_hoc_id = lhp.mon_hoc_id
				JOIN hoc_ky hk ON hk.hoc_ky_id = lhp.hoc_ky_id
				LEFT JOIN ket_qua_hoc_ky kq
				  ON kq.sinh_vien_id = sv.sinh_vien_id
				 AND kq.hoc_ky_id = hk.hoc_ky_id
				WHERE sv.sinh_vien_id = %s
				  AND lhp.trang_thai = 'DA_DUYET'
				ORDER BY hk.ten_hoc_ky DESC, mh.ma_mon ASC
				""",
				(sinh_vien_id,),
			)
			return cursor.fetchall()
	finally:
		conn.close()


def xuat_pdf_bang_diem_ca_nhan(sinh_vien_id: str) -> str:
	try:
		from reportlab.lib.pagesizes import A4
		from reportlab.pdfgen import canvas
	except Exception as exc:  # pragma: no cover
		raise RuntimeError("Thieu thu vien reportlab. Vui long cai dat reportlab.") from exc

	rows = _lay_bang_diem_sinh_vien(sinh_vien_id)
	if not rows:
		raise ValueError("Khong tim thay bang diem sinh vien")

	file_path = _temp_file(".pdf")
	c = canvas.Canvas(file_path, pagesize=A4)
	width, height = A4

	y = height - 40
	first = rows[0]
	c.setFont("Helvetica-Bold", 14)
	c.drawString(40, y, "BANG DIEM CA NHAN")
	y -= 24
	c.setFont("Helvetica", 11)
	c.drawString(40, y, f"MSV: {first['msv']}")
	c.drawString(220, y, f"Ho ten: {first['ho_ten']}")
	y -= 24

	c.setFont("Helvetica-Bold", 10)
	c.drawString(40, y, "Hoc ky")
	c.drawString(120, y, "Mon")
	c.drawString(250, y, "So TC")
	c.drawString(320, y, "Diem tong")
	c.drawString(400, y, "GPA HK")
	c.drawString(470, y, "GPA TL")
	y -= 18
	c.setFont("Helvetica", 10)

	for row in rows:
		if y < 40:
			c.showPage()
			y = height - 40
			c.setFont("Helvetica", 10)

		c.drawString(40, y, str(row.get("ten_hoc_ky") or ""))
		c.drawString(120, y, str(row.get("ma_mon") or ""))
		c.drawString(250, y, str(row.get("so_tin_chi") or ""))
		c.drawString(320, y, str(row.get("diem_tong") or ""))
		c.drawString(400, y, str(row.get("gpa_hk_he4") or ""))
		c.drawString(470, y, str(row.get("gpa_tich_luy_he4") or ""))
		y -= 16

	last = rows[0]
	y -= 8
	c.setFont("Helvetica-Bold", 11)
	c.drawString(40, y, f"Xep loai hien tai: {last.get('xep_loai') or 'Chua xep loai'}")

	c.save()
	return file_path
